import os
import sys
import logging
import requests
from dateutil import parser as dp
import pytz
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.drawing.image import Image
from email.message import EmailMessage
import smtplib
from typing import Optional, Dict, Any

# ---------- Config ----------
BASE_DIR = Path(__file__).parent
ENV_PATH = BASE_DIR / '.env'
load_dotenv(dotenv_path=ENV_PATH)

SAMSARA_TOKEN = os.getenv('SAMSARA_TOKEN')  # Debe iniciar con 'Bearer '
SMTP_HOST = os.getenv('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', 587))
SMTP_USER = os.getenv('SMTP_USER', '')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD', '')
MAIL_TO = os.getenv('MAIL_TO', 'mrodriguez@bgcapitalgroup.mx')  # coma-separado

logging.basicConfig(
    filename=str(BASE_DIR / 'EC03.log'),
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] - %(message)s',
)

MX_TZ = "America/Mexico_City"

# IDs de vehículos (coma, sin espacios)
VEHICLE_IDS = "281474994775389,281474994511222,281474994511248"
PARENT_TAG_IDS = ""  # opcional

# IDs de configuraciones (Entrada/Salida de geocerca)
ENTRY_CONFIG_ID = "c07fc4cb-9bcb-43a7-a496-dc97d8bd759e"  # Geofence Entry
EXIT_CONFIG_ID  = "97ddb8b7-fa02-41eb-a521-f414965d8b47"  # Geofence Exit

# Celdas de la plantilla (ajústalas si difieren)
DATE_CELL = "D3"
TIME_CELL = "G3"
TOTAL_CELL = "I3"      # dónde escribir el total de unidades
START_ROW = 8          # primera fila de datos

# ---------- Utils ----------
def normalize_token(raw: str) -> str:
    raw = (raw or "").strip()
    return raw if raw.startswith("Bearer ") else f"Bearer {raw}"

def write_merged_row(ws, row_idx: int, col_start: int, col_end: int, value, border, align):
    """Escribe 'value' en la celda ancla del merge y aplica formato al rango."""
    ws.merge_cells(start_row=row_idx, start_column=col_start, end_row=row_idx, end_column=col_end)
    anchor = ws.cell(row=row_idx, column=col_start, value=value)
    anchor.border = border
    anchor.alignment = align
    for c in range(col_start + 1, col_end + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = border
        cell.alignment = align

def send_mail_with_attachment(subject: str, body: str, filepath: Path):
    """Envía email con el archivo adjunto."""
    if not SMTP_USER or not SMTP_PASSWORD:
        logging.error("SMTP_USER/SMTP_PASSWORD no configurados")
        raise RuntimeError("Faltan credenciales SMTP")

    recipients = [e.strip() for e in MAIL_TO.split(",") if e.strip()]
    if not recipients:
        logging.error("No hay destinatarios en MAIL_TO")
        raise RuntimeError("MAIL_TO vacío")

    msg = EmailMessage()
    msg["From"] = SMTP_USER
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject
    msg.set_content(body)

    with open(filepath, "rb") as f:
        data = f.read()
        msg.add_attachment(
            data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filepath.name,
        )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASSWORD)
        s.send_message(msg)

# ---------- Incidents helpers ----------
def _rfc3339_utc(dt: datetime) -> str:
    """Devuelve RFC3339 en UTC con 'Z'."""
    return dt.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

def _find_geo_condition(conditions: list) -> Optional[dict]:
    """Devuelve el dict de geofenceEntry o geofenceExit si existe."""
    for c in conditions or []:
        det = c.get("details") or {}
        if "geofenceEntry" in det:
            return det["geofenceEntry"]
        if "geofenceExit" in det:
            return det["geofenceExit"]
    return None

def _last_incident_by_vehicle(
    session: requests.Session,
    configuration_id: str,
    hours_back: int = 168,  # 7 días (cubrir fines de semana)
) -> Dict[str, Dict[str, Any]]:
    """
    Para una configuración (entrada o salida), regresa:
      { vehicleId: { "time": dt_mx, "address_id": str, "address_name": str } }
    con el incidente RESUELTO más reciente por vehículo dentro del rango dado.
    """
    tz_mx = pytz.timezone(MX_TZ)
    start_iso = _rfc3339_utc(datetime.now(timezone.utc) - timedelta(hours=hours_back))
    params = {"startTime": start_iso, "configurationIds": configuration_id}

    try:
        r = session.get("https://api.samsara.com/alerts/incidents/stream", params=params, timeout=60)
        r.raise_for_status()
        data = (r.json() or {}).get("data", [])
        logging.info("Incidentes cfg %s: %s", configuration_id, len(data))
    except Exception:
        logging.exception("Error al consultar incidents/stream cfg=%s", configuration_id)
        return {}

    out: Dict[str, Dict[str, Any]] = {}
    for inc in data:
        if not inc.get("isResolved", False):
            continue

        # Preferir resolvedAtTime; fallback: happened/updated
        at_iso = inc.get("resolvedAtTime") or inc.get("happenedAtTime") or inc.get("updatedAtTime")
        if not at_iso:
            continue

        geo = _find_geo_condition(inc.get("conditions") or [])
        if not geo:
            continue

        vehicle = geo.get("vehicle") or {}
        vid = str(vehicle.get("id") or "")
        if not vid:
            continue

        address = geo.get("address") or {}
        addr_id = str(address.get("id") or "")
        addr_name = address.get("name")

        try:
            at_mx = dp.parse(at_iso).astimezone(tz_mx)
        except Exception:
            continue

        # Conservar el más reciente por vehículo
        if (vid not in out) or (at_mx > out[vid]["time"]):
            out[vid] = {"time": at_mx, "address_id": addr_id, "address_name": addr_name}

    return out

def get_last_entry_exit_by_vehicle(session: requests.Session, hours_back: int = 168) -> Dict[str, Dict[str, Any]]:
    """
    Une las dos consultas (entrada/salida):
      { vehicleId: { "entry": {..} | None, "exit": {..} | None } }
    """
    last_entry = _last_incident_by_vehicle(session, ENTRY_CONFIG_ID, hours_back)
    last_exit  = _last_incident_by_vehicle(session, EXIT_CONFIG_ID,  hours_back)

    all_vids = set(last_entry.keys()) | set(last_exit.keys())
    merged: Dict[str, Dict[str, Any]] = {}
    for vid in all_vids:
        merged[vid] = {
            "entry": last_entry.get(vid),
            "exit":  last_exit.get(vid),
        }
    return merged

# ---------- Main ----------
def fetch_samsara_data():
    logging.info("INICIO DE EJECUCION")
    if not SAMSARA_TOKEN:
        logging.error("SAMSARA_TOKEN no definido en %s", ENV_PATH)
        sys.exit(1)

    token = normalize_token(SAMSARA_TOKEN)

    # Sesión
    session = requests.Session()
    session.headers.update({"Accept": "application/json", "Authorization": token})

    # 1) Feed GPS por vehicleIds
    try:
        params = {"types": "gps", "vehicleIds": VEHICLE_IDS}
        if PARENT_TAG_IDS:
            params["parentTagIds"] = PARENT_TAG_IDS

        resp = session.get("https://api.samsara.com/fleet/vehicles/stats/feed",
                           params=params, timeout=60)
        resp.raise_for_status()
        vehicles = (resp.json() or {}).get("data", [])
        logging.info("Vehiculos en feed: %s", len(vehicles))
    except Exception:
        logging.exception("Error llamando al feed de vehículos")
        sys.exit(1)

    # 1.1) Últimos incidentes de entrada/salida por vehículo (últimos 7 días)
    incidents_by_vehicle = get_last_entry_exit_by_vehicle(session, hours_back=168)

    # 2) Procesar
    tz = pytz.timezone(MX_TZ)
    now_mx = datetime.now(tz)
    results = []

    for u in vehicles:
        try:
            gps_list = u.get("gps") or []
            if isinstance(gps_list, dict):
                gps_list = [gps_list]
            gps_list = [f for f in gps_list if f.get("time")]
            if not gps_list:
                continue
            gps_list.sort(key=lambda f: f["time"], reverse=True)
            g = gps_list[0]

            # Datos
            speed = g.get("speedMilesPerHour", 0) or 0
            ecu = g.get("isEcuSpeed", False) or False
            location = (g.get("reverseGeo") or {}).get("formattedLocation")

            # En patio si hay address en el fix
            addr = g.get("address") or {}
            addr_name = (addr.get("name") or "").strip()
            addr_id = str(addr.get("id") or "")
            en_patio = bool(addr)  # dict vacío -> False
            if en_patio and addr_name:
                location = addr_name

            # Reglas de estatus
            if en_patio:
                estatus = "En patio"
            elif speed == 0 and not ecu:
                continue  # apagado → excluir
            elif speed == 0 and ecu:
                estatus = "Detenido"
            else:
                estatus = "Ruta"

            # ID del vehículo (varias rutas por seguridad)
            u_id = str(u.get("id") or g.get("vehicleId") or (u.get("vehicle") or {}).get("id") or "")

            # Determinar último evento (entró/salió) e imprimir HH:MM:SS en MX
            ult_evento = ""
            hora_entrada = ""
            hora_salida = ""

            if estatus == "En patio" and u_id:
                det = incidents_by_vehicle.get(u_id, {})
                ent = det.get("entry")
                sal = det.get("exit")

                # Valida (si es posible) que la geocerca coincida con el address actual
                def ok_match(cand):
                    if not cand:
                        return False
                    # si ambos tienen address_id, comparamos; si no, aceptamos por vehículo
                    return (not addr_id) or (not cand.get("address_id")) or (cand["address_id"] == addr_id)

                if ent and ok_match(ent):
                    hora_entrada = ent["time"].strftime("%H:%M:%S")
                if sal and ok_match(sal):
                    hora_salida = sal["time"].strftime("%H:%M:%S")

                # Último evento por timestamp
                last_type = None
                last_time = None
                if ent and ok_match(ent):
                    last_type, last_time = "Entró", ent["time"]
                if sal and ok_match(sal) and (last_time is None or sal["time"] > last_time):
                    last_type, last_time = "Salió", sal["time"]

                if last_type and last_time:
                    ult_evento = f"{last_type} {last_time.strftime('%H:%M:%S')}"

            results.append({
                "Unidad": u.get("name", "Desconocido"),
                "Estatus": estatus,
                "Ubicacion": location,
                "UltEvento": ult_evento,     # K
                "HoraEntrada": hora_entrada, # L
                "HoraSalida": hora_salida,   # M
            })
        except Exception:
            logging.warning("Error procesando vehículo %s", u.get("name"))

    print(f"Total unidades: {len(results)}")
    for r in results:
        print(f"{r['Unidad']} | {r['Estatus']} | {r['Ubicacion']} | {r['UltEvento']}")

    # 3) Excel: cargar plantilla (o crear simple)
    plantilla_path = BASE_DIR / "assets" / "templates" / "PlantillaEC03.xlsx"
    if plantilla_path.exists():
        wb = load_workbook(filename=plantilla_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "EC03"
        ws["A7"] = "UNIDAD"
        ws["B7"] = "ESTATUS"    # B..C
        ws["D7"] = "UBICACION"  # D..J
        ws["K7"] = "ULT. EVENTO"
        ws["L7"] = "HORA ENTRADA"
        ws["M7"] = "HORA SALIDA"

    # Asegura encabezados si la plantilla no los tiene
    if not ws["K7"].value: ws["K7"].value = "ULT. EVENTO"
    if not ws["L7"].value: ws["L7"].value = "HORA ENTRADA"
    if not ws["M7"].value: ws["M7"].value = "HORA SALIDA"

    # Limpiar merges de datos anteriores
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= START_ROW:
            ws.unmerge_cells(str(mr))

    # Fecha/Hora
    try:
        ws[DATE_CELL] = now_mx.date().isoformat()
        ws[TIME_CELL] = now_mx.strftime("%H:%M:%S")
    except Exception:
        pass

    # Estilos
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center_al = Alignment(horizontal="center", vertical="center", wrap_text=True)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100", bold=True)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    blue_font = Font(color="1F497D", bold=True)

    # 4) Volcar filas (A=Unidad; B–C=Estatus; D–J=Ubicacion; K=UltEvento; L=Entrada; M=Salida)
    for i, row in enumerate(results, start=START_ROW):
        # Unidad (A)
        c_unidad = ws.cell(row=i, column=1, value=row["Unidad"])
        c_unidad.border = border
        c_unidad.alignment = center_al

        # Estatus (merge B..C)
        write_merged_row(ws, i, 2, 3, row["Estatus"], border, center_al)
        rng = range(2, 4)
        if row["Estatus"] == "Ruta":
            for col in rng:
                c = ws.cell(row=i, column=col); c.fill = green_fill; c.font = green_font
        elif row["Estatus"] == "En patio":
            for col in rng:
                c = ws.cell(row=i, column=col); c.fill = blue_fill; c.font = blue_font
        else:
            for col in rng:
                c = ws.cell(row=i, column=col); c.fill = red_fill; c.font = red_font

        # Ubicación (merge D..J)
        write_merged_row(ws, i, 4, 10, row["Ubicacion"], border, center_al)

        # Ult. Evento (K)
        c_ult = ws.cell(row=i, column=11, value=row["UltEvento"])
        c_ult.border = border
        c_ult.alignment = center_al

        # Hora Entrada (L)
        c_ent = ws.cell(row=i, column=12, value=row["HoraEntrada"])
        c_ent.border = border
        c_ent.alignment = center_al

        # Hora Salida (M)
        c_sal = ws.cell(row=i, column=13, value=row["HoraSalida"])
        c_sal.border = border
        c_sal.alignment = center_al

    # Total de unidades
    last_row = START_ROW + len(results) - 1
    if last_row >= START_ROW:
        try:
            ws[TOTAL_CELL] = f"=COUNTA(A{START_ROW}:A{last_row})"
        except Exception:
            pass

    # Ajustes visuales mínimos
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 14

    # Logo (opcional)
    img_path = BASE_DIR / "assets" / "images" / "TDR-LOGO.png"
    if img_path.exists():
        try:
            logo = Image(str(img_path))
            logo.width = 125
            logo.height = 60
            ws.add_image(logo, "A2")
        except Exception:
            logging.warning("No se pudo insertar el logo")

    # Guardar
    ts_str = now_mx.strftime("%Y-%m-%d_%H-%M-%S")
    out_xlsx = BASE_DIR / f"Reporte_EC03_{ts_str}.xlsx"
    wb.save(out_xlsx)
    logging.info("Archivo generado: %s", out_xlsx)
    print("Archivo generado:", out_xlsx)

    # 5) Enviar correo
    try:
        subject = "Reporte de estatus de unidades EC-03"
        body = "Hola,\n\nSe adjunta el reporte de estatus de unidades.\n\nSaludos."
        send_mail_with_attachment(subject, body, out_xlsx)
        logging.info("Correo enviado a: %s", MAIL_TO)
    except Exception:
        logging.exception("Error enviando correo")
        sys.exit(1)

    # eliminar archivo
    try:
        os.remove(str(out_xlsx))
        logging.info(f"Archivo eliminado: {out_xlsx.name}")
    except Exception:
        logging.exception(f"No se pudo eliminar: {out_xlsx.name}")

    logging.info("===> Ejecución finalizada correctamente")
    logging.info("FIN DE EJECUCION")

if __name__ == "__main__":
    try:
        fetch_samsara_data()
    except Exception:
        logging.exception("Fallo inesperado")
        sys.exit(1)
