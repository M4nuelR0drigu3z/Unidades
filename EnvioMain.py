"""Envio configurable de reportes de telemetria Samsara.

La seleccion, contenido y canales se definen en config/reportes.json.
"""
from __future__ import annotations

import argparse
from copy import deepcopy
from datetime import datetime, timedelta
from email.message import EmailMessage
from io import BytesIO
import json
import logging
import os
from pathlib import Path
import re
import smtplib
import sys
import unicodedata
from typing import Any, Iterable
from urllib.parse import quote

from dateutil import parser as dp
from dotenv import load_dotenv
import gspread
from gspread_dataframe import get_as_dataframe
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pytz
import requests

from detenciones import enriquecer_minutos_detenido


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG_PATH = BASE_DIR / "config" / "reportes.json"
DEFAULT_TAG_CATALOG_PATH = BASE_DIR / "config" / "catalogo_etiquetas_samsara.json"
SAMSARA_BASE_URL = "https://api.samsara.com"
DEFAULT_TIMEZONE = "America/Mexico_City"
MAX_GOOGLE_CHAT_CHARS = 30_000
load_dotenv(dotenv_path=BASE_DIR / ".env")

logging.basicConfig(
    filename=str(BASE_DIR / "reporte_logs.log"),
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)


class ConfiguracionError(ValueError):
    """La configuracion no permite construir un reporte seguro."""


def normalizar_texto(valor: Any) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or "").strip())
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return " ".join(texto.casefold().split())


def slug(valor: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", normalizar_texto(valor)).strip("_") or "reporte"


def cargar_configuracion(ruta: Path) -> dict[str, Any]:
    if not ruta.exists():
        raise ConfiguracionError(f"No existe el archivo de configuracion: {ruta}")
    try:
        config = json.loads(ruta.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        raise ConfiguracionError(f"JSON invalido en {ruta}: {error}") from error

    reportes = config.get("reportes")
    if not isinstance(reportes, list) or not reportes:
        raise ConfiguracionError("La configuracion debe contener 'reportes'.")
    nombres: set[str] = set()
    perfiles = config.get("perfiles_filtros") or {}
    for indice, reporte in enumerate(reportes, start=1):
        nombre = str(reporte.get("nombre") or "").strip()
        if not nombre:
            raise ConfiguracionError(f"El reporte #{indice} no tiene nombre.")
        clave = normalizar_texto(nombre)
        if clave in nombres:
            raise ConfiguracionError(f"Nombre de reporte duplicado: {nombre}")
        nombres.add(clave)
        if not (reporte.get("etiquetas") or reporte.get("etiqueta_ids")):
            raise ConfiguracionError(f"El reporte '{nombre}' necesita etiquetas o etiqueta_ids.")
        tipo = reporte.get("tipo_filtro_etiqueta", "tagIds")
        if tipo not in {"tagIds", "parentTagIds"}:
            raise ConfiguracionError(f"'{nombre}': use tagIds o parentTagIds.")
        perfil = reporte.get("perfil_filtros")
        if perfil and perfil not in perfiles:
            raise ConfiguracionError(f"'{nombre}': perfil_filtros inexistente: {perfil}")
        for entrega in reporte.get("entregas") or []:
            canal = str(entrega.get("canal") or "").strip().lower()
            if canal not in {"google_chat", "correo"}:
                raise ConfiguracionError(f"'{nombre}': canal no soportado: {canal or '(vacio)'}")
    return config


def crear_sesion_samsara(token: str) -> requests.Session:
    sesion = requests.Session()
    sesion.headers.update({"Accept": "application/json", "Authorization": f"Bearer {token}"})
    return sesion


def obtener_paginas(
    sesion: requests.Session,
    ruta: str,
    params: dict[str, Any] | None = None,
    timeout: int = 60,
) -> list[dict[str, Any]]:
    """Consume todas las paginas de un endpoint Samsara basado en cursor."""
    url = ruta if ruta.startswith("http") else f"{SAMSARA_BASE_URL}{ruta}"
    base_params = dict(params or {})
    resultado: list[dict[str, Any]] = []
    after: str | None = None
    while True:
        pagina_params = dict(base_params)
        if after:
            pagina_params["after"] = after
        response = sesion.get(url, params=pagina_params, timeout=timeout)
        logging.info("GET %s status=%s", response.url, response.status_code)
        response.raise_for_status()
        payload = response.json()
        resultado.extend(payload.get("data") or [])
        pagination = payload.get("pagination") or {}
        if not pagination.get("hasNextPage"):
            return resultado
        after = pagination.get("endCursor")
        if not after:
            raise RuntimeError(f"{ruta} indico otra pagina pero no envio endCursor")


def listar_etiquetas_samsara(sesion: requests.Session) -> list[dict[str, Any]]:
    return obtener_paginas(sesion, "/tags", {"limit": 512})


def obtener_catalogo_etiquetas_samsara(
    sesion: requests.Session, limit: int = 10
) -> list[dict[str, Any]]:
    """Descarga tags por lotes pequenos y conserva solo datos de jerarquia."""
    resultado: list[dict[str, Any]] = []
    after: str | None = None
    while True:
        params: dict[str, Any] = {"limit": limit}
        if after:
            params["after"] = after
        response = sesion.get(f"{SAMSARA_BASE_URL}/tags", params=params, timeout=120)
        response.raise_for_status()
        payload = response.json()
        for tag in payload.get("data") or []:
            resultado.append(
                {
                    "id": str(tag.get("id") or ""),
                    "nombre": str(tag.get("name") or "").strip(),
                    "parentTagId": str(tag.get("parentTagId") or ""),
                }
            )
        pagination = payload.get("pagination") or {}
        print(f"[Tags] Descargadas: {len(resultado)}", flush=True)
        if not pagination.get("hasNextPage"):
            return resultado
        after = pagination.get("endCursor")
        if not after:
            raise RuntimeError("Samsara no devolvio endCursor para la siguiente pagina de tags")


def construir_jerarquia_etiquetas(tags: list[dict[str, Any]]) -> dict[str, Any]:
    nodos = {
        str(tag["id"]): {
            "id": str(tag["id"]),
            "nombre": str(tag.get("nombre") or ""),
            "parentTagId": str(tag.get("parentTagId") or ""),
            "hijos": [],
        }
        for tag in tags
        if tag.get("id")
    }
    raices = []
    for nodo in nodos.values():
        padre = nodos.get(nodo["parentTagId"])
        if padre:
            padre["hijos"].append(nodo)
        else:
            raices.append(nodo)

    def ordenar(nodo: dict[str, Any]) -> None:
        nodo["hijos"].sort(key=lambda x: normalizar_texto(x["nombre"]))
        for hijo in nodo["hijos"]:
            ordenar(hijo)

    raices.sort(key=lambda x: normalizar_texto(x["nombre"]))
    for raiz in raices:
        ordenar(raiz)
    padres = sorted(
        (
            {"id": nodo["id"], "nombre": nodo["nombre"], "parentTagId": nodo["parentTagId"]}
            for nodo in nodos.values()
            if nodo["hijos"]
        ),
        key=lambda x: normalizar_texto(x["nombre"]),
    )
    return {
        "generado": datetime.now(pytz.timezone(DEFAULT_TIMEZONE)).isoformat(),
        "totalEtiquetas": len(nodos),
        "totalEtiquetasPadre": len(padres),
        "etiquetasPadre": padres,
        "jerarquia": raices,
        "etiquetas": sorted(tags, key=lambda x: normalizar_texto(x.get("nombre"))),
    }


def guardar_catalogo_etiquetas(
    sesion: requests.Session, ruta: Path, limit: int = 10
) -> dict[str, Any]:
    catalogo = construir_jerarquia_etiquetas(obtener_catalogo_etiquetas_samsara(sesion, limit))
    ruta.parent.mkdir(parents=True, exist_ok=True)
    ruta.write_text(json.dumps(catalogo, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return catalogo


def cargar_cache_catalogo(ruta: Path) -> dict[str, dict[str, str]]:
    if not ruta.exists():
        return {}
    data = json.loads(ruta.read_text(encoding="utf-8"))
    return {
        normalizar_texto(tag.get("nombre")): {
            "id": str(tag.get("id") or ""),
            "name": str(tag.get("nombre") or ""),
        }
        for tag in data.get("etiquetas") or []
        if tag.get("id") and tag.get("nombre")
    }


def resolver_etiquetas(
    nombres_solicitados: Iterable[str], etiquetas_samsara: Iterable[dict[str, Any]]
) -> dict[str, str]:
    """Resuelve nombres exactos ignorando mayusculas, espacios y acentos."""
    indice: dict[str, list[dict[str, Any]]] = {}
    for etiqueta in etiquetas_samsara:
        nombre = str(etiqueta.get("name") or "").strip()
        etiqueta_id = str(etiqueta.get("id") or "").strip()
        if nombre and etiqueta_id:
            indice.setdefault(normalizar_texto(nombre), []).append(etiqueta)
    encontrados: dict[str, str] = {}
    faltantes, ambiguas = [], []
    for solicitado in nombres_solicitados:
        coincidencias = indice.get(normalizar_texto(solicitado), [])
        if not coincidencias:
            faltantes.append(str(solicitado))
        elif len(coincidencias) > 1:
            ambiguas.append(str(solicitado))
        else:
            encontrados[str(solicitado)] = str(coincidencias[0]["id"])
    if faltantes or ambiguas:
        partes = []
        if faltantes:
            partes.append("no encontradas: " + ", ".join(faltantes))
        if ambiguas:
            partes.append("duplicadas en Samsara: " + ", ".join(ambiguas))
        raise ConfiguracionError("Etiquetas " + "; ".join(partes))
    return encontrados


def obtener_ids_etiquetas_reporte(
    reporte: dict[str, Any], etiquetas_samsara: list[dict[str, Any]]
) -> tuple[list[str], dict[str, str]]:
    resueltas = resolver_etiquetas(reporte.get("etiquetas") or [], etiquetas_samsara)
    ids = [str(x).strip() for x in reporte.get("etiqueta_ids") or [] if str(x).strip()]
    ids.extend(resueltas.values())
    return list(dict.fromkeys(ids)), resueltas


def resolver_etiquetas_samsara(
    sesion: requests.Session,
    nombres_solicitados: Iterable[str],
    cache: dict[str, dict[str, str]] | None = None,
) -> dict[str, str]:
    """Busca cada tag por el external ID automatico ``samsara.name``."""
    cache = cache if cache is not None else {}
    resultado: dict[str, str] = {}
    for nombre in nombres_solicitados:
        clave = normalizar_texto(nombre)
        if clave not in cache:
            external_id = quote(f"samsara.name:{str(nombre).strip()}", safe=":")
            response = sesion.get(f"{SAMSARA_BASE_URL}/tags/{external_id}", timeout=60)
            if response.status_code == 404:
                raise ConfiguracionError(f"Etiqueta no encontrada en Samsara: {nombre}")
            response.raise_for_status()
            data = response.json().get("data") or {}
            etiqueta_id = str(data.get("id") or "").strip()
            nombre_real = str(data.get("name") or "").strip()
            if not etiqueta_id or normalizar_texto(nombre_real) != clave:
                raise ConfiguracionError(
                    f"Samsara no devolvio una coincidencia exacta para: {nombre}"
                )
            cache[clave] = {"id": etiqueta_id, "name": nombre_real}
        resultado[str(nombre)] = cache[clave]["id"]
    return resultado


def obtener_ids_etiquetas_reporte_remoto(
    sesion: requests.Session,
    reporte: dict[str, Any],
    cache: dict[str, dict[str, str]],
) -> tuple[list[str], dict[str, str]]:
    resueltas = resolver_etiquetas_samsara(sesion, reporte.get("etiquetas") or [], cache)
    ids = [str(x).strip() for x in reporte.get("etiqueta_ids") or [] if str(x).strip()]
    ids.extend(resueltas.values())
    return list(dict.fromkeys(ids)), resueltas


def obtener_geocercas_de_etiquetas(
    sesion: requests.Session, etiqueta_ids: Iterable[str]
) -> dict[str, str]:
    geocercas: dict[str, str] = {}
    for etiqueta_id in dict.fromkeys(str(x) for x in etiqueta_ids if x):
        response = sesion.get(f"{SAMSARA_BASE_URL}/tags/{etiqueta_id}", timeout=60)
        response.raise_for_status()
        data = response.json().get("data") or {}
        for address in data.get("addresses") or []:
            address_id = str(address.get("id") or "").strip()
            if address_id:
                geocercas[address_id] = str(address.get("name") or "").strip()
    return geocercas


def obtener_vehiculos(
    sesion: requests.Session, etiqueta_ids: list[str], tipo_filtro: str
) -> list[dict[str, Any]]:
    return obtener_paginas(
        sesion,
        "/fleet/vehicles/stats",
        {"types": "gps,engineStates,ecuSpeedMph", tipo_filtro: ",".join(etiqueta_ids)},
    )


def enriquecer_operadores_samsara(
    sesion: requests.Session,
    results: list[dict[str, Any]],
    now_mx: datetime,
    ventana_horas: int = 24,
) -> list[dict[str, Any]]:
    """Asigna el operador no pasajero mas reciente reportado por Samsara."""
    vehiculo_ids = [str(row.get("SamsaraVehicleId") or "") for row in results]
    vehiculo_ids = list(dict.fromkeys(x for x in vehiculo_ids if x))
    for row in results:
        row.setdefault("Operador", "")
        row.setdefault("ID Operador", "")
    if not vehiculo_ids:
        return results

    fin = now_mx.astimezone(pytz.UTC)
    inicio = fin - timedelta(hours=max(1, int(ventana_horas)))
    asignaciones = obtener_paginas(
        sesion,
        "/fleet/driver-vehicle-assignments",
        {
            "filterBy": "vehicles",
            "vehicleIds": ",".join(vehiculo_ids),
            "startTime": inicio.isoformat(),
            "endTime": fin.isoformat(),
        },
    )
    recientes: dict[str, tuple[datetime, dict[str, Any]]] = {}
    for asignacion in asignaciones:
        if asignacion.get("isPassenger"):
            continue
        vehiculo_id = str((asignacion.get("vehicle") or {}).get("id") or "")
        conductor = asignacion.get("driver") or {}
        inicio_asignacion = asignacion.get("startTime")
        if not vehiculo_id or not conductor.get("name") or not inicio_asignacion:
            continue
        fecha = dp.parse(inicio_asignacion)
        if vehiculo_id not in recientes or fecha > recientes[vehiculo_id][0]:
            recientes[vehiculo_id] = (fecha, conductor)

    encontrados = 0
    for row in results:
        asignacion = recientes.get(str(row.get("SamsaraVehicleId") or ""))
        if not asignacion:
            continue
        conductor = asignacion[1]
        row["Operador"] = str(conductor.get("name") or "")
        row["ID Operador"] = str(conductor.get("id") or "")
        encontrados += 1
    print(f"[Samsara] Operadores vinculados: {encontrados}/{len(results)}")
    return results


def obtener_datos_google_sheets(
    results: list[dict[str, Any]], fecha_busqueda: datetime, settings: dict[str, Any]
) -> list[dict[str, Any]]:
    """Enriquece placas, identificador y operador desde la planeacion diaria."""
    for row in results:
        row.setdefault("ID ROSTERING", "")
        row.setdefault("Operador", "")
        row.setdefault("PLACAS", "")
        row.setdefault("ORIGEN", "")
        row.setdefault("DESTINO", "")
    if not results:
        return results

    credentials = BASE_DIR / settings.get("credenciales", "Credenciales.json")
    sheet_url = settings.get(
        "url",
        "https://docs.google.com/spreadsheets/d/1zbVe4Rk7aGaC_gyy0n2ik5VEWzn3w6XAyn91LNp2cMA/edit#gid=0",
    )
    gc = gspread.service_account(filename=credentials)
    workbook = gc.open_by_url(sheet_url)
    meses = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
    }
    candidatos = [
        f"{meses[fecha_busqueda.month]} {str(fecha_busqueda.year)[-2:]}",
        f"{meses[fecha_busqueda.month].title()} {fecha_busqueda.year}",
    ]
    hojas = {normalizar_texto(ws.title): ws.title for ws in workbook.worksheets()}
    nombre_hoja = next(
        (hojas[normalizar_texto(x)] for x in candidatos if normalizar_texto(x) in hojas),
        None,
    )
    if not nombre_hoja:
        logging.warning("No se encontro hoja mensual. Candidatas: %s", candidatos)
        return results

    df = get_as_dataframe(workbook.worksheet(nombre_hoja), evaluate_formulas=True).fillna("")
    indice_columnas = {normalizar_texto(columna): columna for columna in df.columns}

    def encontrar_columna(*candidatos: str) -> str | None:
        return next(
            (indice_columnas[normalizar_texto(x)] for x in candidatos if normalizar_texto(x) in indice_columnas),
            None,
        )

    columna_fecha = encontrar_columna("FECHA DE INICIO", "FECHA")
    columna_unidad = encontrar_columna("UNIDAD")
    columna_operador = encontrar_columna("OPERADOR", "OP")
    columna_roster = encontrar_columna("ROSTERING ID", "ROSTERING \nID", "ID OP 1", "ROSTER")
    columna_placas = encontrar_columna("PLACAS")
    if not columna_fecha or not columna_unidad:
        logging.warning("Faltan columnas de fecha o unidad en la hoja %s", nombre_hoja)
        return results

    meses_es_en = {
        "ene": "jan", "feb": "feb", "mar": "mar", "abr": "apr",
        "may": "may", "jun": "jun", "jul": "jul", "ago": "aug",
        "sep": "sep", "sept": "sep", "set": "sep", "oct": "oct",
        "nov": "nov", "dic": "dec",
    }

    def normalizar_fecha(valor: Any):
        texto = str(valor).strip().lower()
        for es, en in meses_es_en.items():
            texto = re.sub(rf"\b{es}\b", en, texto)
        try:
            return dp.parse(texto, dayfirst=True).date()
        except Exception:
            return None

    df["_fecha_norm"] = df[columna_fecha].apply(normalizar_fecha)
    filas = df[df["_fecha_norm"] == fecha_busqueda.date()]
    operadores_encontrados = 0
    for row in results:
        unidad = str(row.get("Unidad") or "").strip()
        coincidencias = filas[filas[columna_unidad].astype(str).str.strip() == unidad]
        if coincidencias.empty:
            continue
        if columna_operador:
            con_operador = coincidencias[
                coincidencias[columna_operador].astype(str).str.strip() != ""
            ]
            fila = (con_operador if not con_operador.empty else coincidencias).iloc[-1]
            row["Operador"] = fila.get(columna_operador, "")
            if str(row["Operador"]).strip():
                operadores_encontrados += 1
        else:
            fila = coincidencias.iloc[-1]
        row["ID ROSTERING"] = fila.get(columna_roster, "") if columna_roster else ""
        row["PLACAS"] = fila.get(columna_placas, "") if columna_placas else ""
    logging.info("Operadores vinculados desde Sheets: %s/%s", operadores_encontrados, len(results))
    print(f"[Sheets] Operadores vinculados: {operadores_encontrados}/{len(results)}")
    return results


def procesar_vehiculos(
    vehicles: Iterable[dict[str, Any]],
    now_mx: datetime,
    filtros: dict[str, Any],
    geocercas_excluidas: dict[str, str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    results, excluidas = [], []
    gps_max_minutos = filtros.get("gps_max_minutos", 60)
    incluir_todas = filtros.get("incluir_todas_las_unidades", False)
    especiales = {str(x) for x in filtros.get("geocercas_especiales_ids") or []}
    for u in vehicles:
        unidad_id = str(u.get("id") or "")
        unidad = str(u.get("name") or "Sin nombre")
        try:
            gps = u.get("gps") or {}
            gps_time = gps.get("time")
            loc_time = dp.parse(gps_time).astimezone(now_mx.tzinfo) if gps_time else None
            antiguedad = int((now_mx - loc_time).total_seconds() / 60) if loc_time else None
            address = gps.get("address") or {}
            geocerca_id = str(address.get("id") or "")
            geocerca = str(address.get("name") or "").strip()
            speed = float(gps.get("speedMilesPerHour") or 0)
            ecu = bool(gps.get("isEcuSpeed", False))
            lat, lon = gps.get("latitude", ""), gps.get("longitude", "")
            coordenadas = f"{lat},{lon}" if lat != "" and lon != "" else ""
            motivo, detalle = "", ""
            if gps_max_minutos is not None and (antiguedad is None or antiguedad > int(gps_max_minutos)):
                motivo = "GPS SIN FECHA" if loc_time is None else "GPS VIEJO"
                detalle = "GPS sin marca de tiempo" if loc_time is None else f"Antiguedad: {antiguedad} minutos"
            elif geocerca_id in especiales:
                motivo, detalle = "GEOCERCA ESPECIAL", "El ID esta configurado como geocerca especial"
            elif geocerca_id in geocercas_excluidas:
                motivo, detalle = "PATIO/GEOCERCA EXCLUIDA", "La geocerca pertenece a una etiqueta excluida"
                geocerca = geocercas_excluidas[geocerca_id] or geocerca
            elif filtros.get("excluir_speed_cero_sin_ecu", True) and speed == 0 and not ecu:
                motivo, detalle = "SPEED 0 SIN ECU", "speedMilesPerHour es 0 e isEcuSpeed es False"
            if motivo and not incluir_todas:
                excluidas.append({
                    "Unidad": unidad, "SamsaraVehicleId": unidad_id, "Motivo": motivo,
                    "Detalle": detalle, "GpsTimeMexico": loc_time,
                    "AntiguedadMinutos": antiguedad, "GeocercaId": geocerca_id,
                    "Geocerca": geocerca, "Latitud": lat, "Longitud": lon,
                    "Coordenadas": coordenadas, "VelocidadMph": speed, "IsEcuSpeed": ecu,
                })
                continue
            location = (gps.get("reverseGeo") or {}).get("formattedLocation", "")
            results.append({
                "Unidad": unidad, "SamsaraVehicleId": unidad_id, "GpsActual": gps,
                "Fecha GPS": loc_time, "Ubicación": location,
                "Estatus": "DETENIDO" if speed == 0 and ecu else "RUTA",
                "Velocidad Mph": speed, "IsEcuSpeed": ecu,
                "Latitud": lat, "Longitud": lon, "Coordenadas": coordenadas,
                "Geocerca": geocerca,
                "Minutos Detenido": None, "Tiempo Detenido": None,
                "Detenido Desde": None, "Ventana Detenido": "",
                "Minutos Trafico": None, "Tiempo Trafico": None,
                "Trafico Desde": None, "Motor": "", "EcuSpeedActual": None,
            })
        except Exception as error:
            logging.exception("Error procesando unidad %s", unidad)
            if incluir_todas:
                results.append({
                    "Unidad": unidad, "SamsaraVehicleId": unidad_id,
                    "Operador": "", "Ubicación": "", "Coordenadas": "",
                })
            else:
                excluidas.append({"Unidad": unidad, "SamsaraVehicleId": unidad_id, "Motivo": "ERROR", "Detalle": str(error)})
    return results, excluidas


def filtrar_contenido(results: list[dict[str, Any]], contenido: dict[str, Any]):
    estados = {normalizar_texto(x) for x in contenido.get("estados") or []}
    unidades = {normalizar_texto(x) for x in contenido.get("unidades") or []}
    filtrados = [
        row for row in results
        if (not estados or normalizar_texto(row.get("Estatus")) in estados)
        and (not unidades or normalizar_texto(row.get("Unidad")) in unidades)
    ]
    orden_estados = {
        normalizar_texto(estado): indice
        for indice, estado in enumerate(contenido.get("orden_estados") or [])
    }
    if orden_estados:
        def clave_unidad(row: dict[str, Any]):
            unidad = str(row.get("Unidad") or "")
            return (0, int(unidad)) if unidad.isdigit() else (1, normalizar_texto(unidad))

        filtrados.sort(key=lambda row: (
            orden_estados.get(normalizar_texto(row.get("Estatus")), len(orden_estados)),
            clave_unidad(row),
        ))
    return filtrados


def construir_reporte_google(results: list[dict[str, Any]], now_mx: datetime, titulo: str):
    conteos: dict[str, int] = {}
    for row in results:
        estado = str(row.get("Estatus") or "SIN ESTATUS")
        conteos[estado] = conteos.get(estado, 0) + 1
    lineas = [
        f"🚛 *{titulo}*", "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
        f"📅 *Fecha:* {now_mx:%Y-%m-%d}", f"🕒 *Hora:* {now_mx:%H:%M:%S}",
        f"📦 *Total unidades:* {len(results)}", "", "📊 *Resumen*",
    ]
    iconos = {"RUTA": "✅", "DETENIDO": "⛔", "TRAFICO LENTO": "🚦", "RETEN": "🚧"}
    for estado in iconos:
        lineas.append(f"{iconos[estado]} {estado.title()}: {conteos.get(estado, 0)}")
    lineas.extend(["", "*Detalle:*", "```"])
    if not results:
        lineas.append("No se encontraron unidades para reportar.")
    else:
        lineas.append(f"{'UNIDAD':<12} | {'ESTATUS':<14} | {'TIEMPO':<12} | {'COORDENADAS':<23} | UBICACION")
        lineas.append("-" * 125)
        orden = {"DETENIDO": 0, "TRAFICO LENTO": 1, "RETEN": 2, "RUTA": 3}
        for row in sorted(results, key=lambda x: (orden.get(x.get("Estatus"), 99), str(x.get("Unidad")))):
            estado = str(row.get("Estatus") or "")
            tiempo = row.get("Tiempo Detenido") if estado == "DETENIDO" else row.get("Tiempo Trafico")
            ubicacion = str(row.get("Ubicación") or "").replace("\n", " ")[:62]
            coordenadas = str(row.get("Coordenadas") or "")[:23]
            lineas.append(f"{str(row.get('Unidad') or '')[:12]:<12} | {estado[:14]:<14} | {str(tiempo or '')[:12]:<12} | {coordenadas:<23} | {ubicacion}")
    lineas.extend(["```", "", "✅ *Reporte generado automáticamente*"])
    return "\n".join(lineas)


def dividir_mensaje(texto: str, max_chars: int = MAX_GOOGLE_CHAT_CHARS) -> list[str]:
    if len(texto) <= max_chars:
        return [texto]
    partes, actual, longitud = [], [], 0
    for linea in texto.splitlines():
        costo = len(linea) + 1
        if actual and longitud + costo > max_chars:
            partes.append("\n".join(actual)); actual, longitud = [], 0
        actual.append(linea); longitud += costo
    if actual:
        partes.append("\n".join(actual))
    return partes


def enviar_google_chat(texto: str, webhook_url: str, sesion=requests) -> None:
    for parte in dividir_mensaje(texto):
        response = sesion.post(webhook_url, json={"text": parte}, timeout=30)
        response.raise_for_status()


DEFAULT_COLUMNS = [
    "Unidad", "Estatus", "Tiempo Detenido", "Tiempo Trafico", "Motor",
    "Fecha GPS", "Ubicación", "Coordenadas", "Geocerca", "Velocidad Mph",
    "IsEcuSpeed", "PLACAS", "ID ROSTERING",
]


def normalizar_valor_excel(valor: Any) -> Any:
    if isinstance(valor, datetime):
        return valor.replace(tzinfo=None)
    return valor if isinstance(valor, (str, int, float, bool)) or valor is None else str(valor)


def aplicar_estilo_hoja(ws, fila_encabezado: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{fila_encabezado + 1}"
    for cell in ws[fila_encabezado]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for columna in range(1, ws.max_column + 1):
        max_len = max((len(str(ws.cell(fila, columna).value or "")) for fila in range(1, ws.max_row + 1)), default=0)
        ws.column_dimensions[get_column_letter(columna)].width = min(max(max_len + 2, 12), 55)
    if ws.max_row > fila_encabezado:
        ws.auto_filter.ref = f"A{fila_encabezado}:{get_column_letter(ws.max_column)}{ws.max_row}"
        for row in ws.iter_rows(min_row=fila_encabezado + 1):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="D9EAF7")


def crear_excel_reporte(
    nombre: str, results: list[dict[str, Any]], excluidas: list[dict[str, Any]],
    now_mx: datetime, contenido: dict[str, Any],
) -> bytes:
    wb = Workbook()
    ws = wb.active; ws.title = "Unidades"
    columnas = contenido.get("columnas") or DEFAULT_COLUMNS
    ws.append([nombre]); ws.append(["Generado", now_mx.replace(tzinfo=None)])
    if contenido.get("resumen_estados_excel", False):
        conteos = {
            estado: sum(1 for row in results if normalizar_texto(row.get("Estatus")) == normalizar_texto(estado))
            for estado in ("RUTA", "DETENIDO")
        }
        ws.append(["Total", len(results), "En ruta", conteos["RUTA"], "Detenidos", conteos["DETENIDO"]])
        for referencia in ("A3", "C3", "E3"):
            ws[referencia].font = Font(bold=True, color="17365D")
    else:
        ws.append(["Total", len(results)])
    ws.append([]); ws.append(columnas)
    for row in results:
        ws.append([normalizar_valor_excel(row.get(columna)) for columna in columnas])
    aplicar_estilo_hoja(ws, 5); ws["B2"].number_format = "yyyy-mm-dd hh:mm:ss"

    if contenido.get("incluir_resumen_excel", True):
        resumen = wb.create_sheet("Resumen", 0)
        resumen.append(["Resumen del reporte", nombre]); resumen.append(["Generado", now_mx.replace(tzinfo=None)])
        resumen.append(["Unidades incluidas", len(results)]); resumen.append(["Unidades omitidas", len(excluidas)])
        estados: dict[str, int] = {}
        for row in results:
            estado = str(row.get("Estatus") or "SIN ESTATUS"); estados[estado] = estados.get(estado, 0) + 1
        for estado, cantidad in sorted(estados.items()):
            resumen.append([estado, cantidad])
        aplicar_estilo_hoja(resumen, 1); resumen["B2"].number_format = "yyyy-mm-dd hh:mm:ss"

    if contenido.get("incluir_omitidas_excel", True):
        omitidas_ws = wb.create_sheet("Omitidas")
        columnas_omitidas = ["Unidad", "SamsaraVehicleId", "Motivo", "Detalle", "GpsTimeMexico", "AntiguedadMinutos", "Geocerca", "GeocercaId", "Latitud", "Longitud", "Coordenadas", "VelocidadMph", "IsEcuSpeed"]
        omitidas_ws.append(columnas_omitidas)
        for row in excluidas:
            omitidas_ws.append([normalizar_valor_excel(row.get(c)) for c in columnas_omitidas])
        aplicar_estilo_hoja(omitidas_ws, 1)
    salida = BytesIO(); wb.save(salida)
    return salida.getvalue()


def obtener_destinatarios(entrega: dict[str, Any]) -> list[str]:
    destinos = [str(x).strip() for x in entrega.get("destinatarios") or [] if str(x).strip()]
    env_name = str(entrega.get("destinatarios_env") or "").strip()
    if env_name:
        destinos.extend(x.strip() for x in os.getenv(env_name, "").split(",") if x.strip())
    return list(dict.fromkeys(destinos))


def enviar_correo_excel(destinatarios: list[str], asunto: str, cuerpo: str, nombre_archivo: str, excel: bytes) -> None:
    host, port = os.getenv("SMTP_HOST", "smtp.gmail.com"), int(os.getenv("SMTP_PORT", "587"))
    user, password = os.getenv("SMTP_USER", ""), os.getenv("SMTP_PASSWORD", "")
    from_address = os.getenv("EMAIL_FROM_ADDRESS", "") or user
    from_name = os.getenv("EMAIL_FROM_NAME", "Reportes de telemetria")
    if not user or not password or not from_address:
        raise ConfiguracionError("Faltan SMTP_USER/SMTP_PASSWORD/EMAIL_FROM_ADDRESS")
    if not destinatarios:
        raise ConfiguracionError("La entrega por correo no tiene destinatarios")
    msg = EmailMessage(); msg["From"] = f"{from_name} <{from_address}>"
    msg["To"] = ", ".join(destinatarios); msg["Subject"] = asunto; msg.set_content(cuerpo)
    msg.add_attachment(excel, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=nombre_archivo)
    smtp_cls = smtplib.SMTP_SSL if port == 465 else smtplib.SMTP
    with smtp_cls(host, port, timeout=60) as smtp:
        if port != 465:
            smtp.starttls()
        smtp.login(user, password); smtp.send_message(msg)


def formatear_plantilla(texto: str, reporte: dict[str, Any], now_mx: datetime) -> str:
    return texto.format(nombre=reporte["nombre"], equipo=reporte.get("equipo", reporte["nombre"]), fecha=now_mx.strftime("%Y-%m-%d"), hora=now_mx.strftime("%H:%M:%S"))


def seleccionar_entregas(reporte: dict[str, Any], canal_forzado: str | None = None) -> list[dict[str, Any]]:
    entregas = reporte.get("entregas") or []
    if canal_forzado:
        seleccionadas = [
            entrega for entrega in entregas
            if str(entrega.get("canal", "")).lower() == canal_forzado.lower()
        ]
        if not seleccionadas:
            raise ConfiguracionError(
                f"El reporte {reporte['nombre']} no tiene configurado el canal {canal_forzado}"
            )
        return seleccionadas
    return [entrega for entrega in entregas if entrega.get("activo", True)]


def entregar_reporte(reporte, results, excluidas, now_mx, dry_run, canal_forzado=None):
    contenido = reporte.get("contenido") or {}
    entregas = seleccionar_entregas(reporte, canal_forzado)
    if not entregas:
        print(f"[{reporte['nombre']}] Sin entregas activas; no se envio nada."); return
    excel_cache = None
    for entrega in entregas:
        canal = entrega["canal"].lower()
        if canal == "google_chat":
            mensaje = construir_reporte_google(results, now_mx, reporte["nombre"])
            if dry_run:
                print(f"\n[DRY-RUN][Google Chat][{reporte['nombre']}]\n{mensaje}"); continue
            env_name = entrega.get("webhook_env", "GOOGLE_CHAT_WEBHOOK_URL")
            webhook = os.getenv(env_name, "")
            if not webhook:
                raise ConfiguracionError(f"Falta la variable {env_name}")
            enviar_google_chat(mensaje, webhook)
            print(f"[{reporte['nombre']}] Enviado a Google Chat.")
        elif canal == "correo":
            if excel_cache is None:
                excel_cache = crear_excel_reporte(reporte["nombre"], results, excluidas, now_mx, contenido)
            filename = f"{slug(reporte['nombre'])}_{now_mx:%Y%m%d_%H%M%S}.xlsx"
            if dry_run:
                destino = BASE_DIR / "outputs" / "envio_previews" / filename
                destino.parent.mkdir(parents=True, exist_ok=True); destino.write_bytes(excel_cache)
                load_workbook(BytesIO(excel_cache), read_only=True).close()
                print(f"[DRY-RUN][Correo] Excel generado: {destino}"); continue
            destinatarios = obtener_destinatarios(entrega)
            asunto = formatear_plantilla(entrega.get("asunto", "{nombre} - {fecha} {hora}"), reporte, now_mx)
            cuerpo = formatear_plantilla(entrega.get("cuerpo", "Se adjunta el reporte {nombre}."), reporte, now_mx)
            enviar_correo_excel(destinatarios, asunto, cuerpo, filename, excel_cache)
            print(f"[{reporte['nombre']}] Enviado por correo a {len(destinatarios)} destinatario(s).")


def combinar_filtros(base, reporte, perfiles=None):
    resultado = deepcopy(base)
    perfil = reporte.get("perfil_filtros")
    if perfil:
        resultado.update(deepcopy((perfiles or {}).get(perfil) or {}))
    resultado.update(reporte.get("filtros") or {})
    return resultado


def seleccionar_reportes(config, nombres):
    activos = [r for r in config["reportes"] if r.get("activo", True)]
    if not nombres:
        return activos
    solicitados = {normalizar_texto(x) for x in nombres}
    seleccionados = [
        r for r in config["reportes"]
        if normalizar_texto(r["nombre"]) in solicitados
    ]
    faltantes = solicitados - {normalizar_texto(r["nombre"]) for r in seleccionados}
    if faltantes:
        raise ConfiguracionError("Reportes no encontrados: " + ", ".join(sorted(faltantes)))
    return seleccionados


def ejecutar(args: argparse.Namespace) -> int:
    if not 1 <= args.tag_page_size <= 512:
        raise ConfiguracionError("--tag-page-size debe estar entre 1 y 512")
    config = cargar_configuracion(args.config)
    reportes = seleccionar_reportes(config, args.solo)
    token = os.getenv("SAMSARA_API_TOKEN") or os.getenv("SAMSARA_TOKEN", "").removeprefix("Bearer ")
    if not token:
        raise ConfiguracionError("Falta SAMSARA_API_TOKEN o SAMSARA_TOKEN")
    sesion = crear_sesion_samsara(token)
    catalogo_path = Path(config.get("catalogo_etiquetas", DEFAULT_TAG_CATALOG_PATH))
    if not catalogo_path.is_absolute():
        catalogo_path = BASE_DIR / catalogo_path
    if args.sincronizar_etiquetas:
        catalogo = guardar_catalogo_etiquetas(sesion, catalogo_path, args.tag_page_size)
        print(
            f"Catalogo generado: {catalogo_path} | "
            f"Etiquetas={catalogo['totalEtiquetas']} Padres={catalogo['totalEtiquetasPadre']}"
        )
        return 0
    if args.listar_etiquetas:
        tags = obtener_catalogo_etiquetas_samsara(sesion, args.tag_page_size)
        for tag in sorted(tags, key=lambda x: normalizar_texto(x.get("nombre"))):
            print(f"{tag.get('id')}\t{tag.get('parentTagId')}\t{tag.get('nombre')}")
        return 0
    tz = pytz.timezone(config.get("zona_horaria", DEFAULT_TIMEZONE)); now_mx = datetime.now(tz)
    filtros_base = config.get("filtros_base") or {}
    cache_geocercas: dict[tuple[str, ...], dict[str, str]] = {}
    cache_etiquetas = cargar_cache_catalogo(catalogo_path)
    errores = []
    for reporte in reportes:
        try:
            ids, resueltas = obtener_ids_etiquetas_reporte_remoto(
                sesion, reporte, cache_etiquetas
            )
            print(f"\n[{reporte['nombre']}] Etiquetas: {resueltas or ids}")
            vehicles = obtener_vehiculos(sesion, ids, reporte.get("tipo_filtro_etiqueta", "tagIds"))
            filtros = combinar_filtros(
                filtros_base, reporte, config.get("perfiles_filtros") or {}
            )
            excluir_geocercas = filtros.get("excluir_unidades_en_geocerca", False)
            ids_geocercas = []
            if excluir_geocercas:
                geocercas_por_nombre = resolver_etiquetas_samsara(
                    sesion, filtros.get("etiquetas_geocercas_excluidas") or [], cache_etiquetas
                )
                ids_geocercas.extend(geocercas_por_nombre.values())
                ids_geocercas.extend(
                    str(x) for x in filtros.get("etiqueta_ids_geocercas_excluidas") or []
                )
            ids_geocercas = tuple(dict.fromkeys(ids_geocercas))
            if ids_geocercas not in cache_geocercas:
                cache_geocercas[ids_geocercas] = obtener_geocercas_de_etiquetas(sesion, ids_geocercas)
            results, excluidas = procesar_vehiculos(
                vehicles, now_mx, filtros, cache_geocercas[ids_geocercas]
            )
            sheets_settings = config.get("google_sheets") or {}
            if sheets_settings.get("activo", True) and reporte.get("enriquecer_google_sheets", True):
                try:
                    results = obtener_datos_google_sheets(results, now_mx, sheets_settings)
                except Exception:
                    logging.exception("No se pudo enriquecer desde Google Sheets")
            if reporte.get("enriquecer_operadores_samsara", False):
                try:
                    results = enriquecer_operadores_samsara(
                        sesion,
                        results,
                        now_mx,
                        reporte.get("operadores_ventana_horas", 24),
                    )
                except Exception:
                    logging.exception("No se pudo enriquecer operadores desde Samsara")
            if reporte.get("analizar_detenciones", True):
                try:
                    results = enriquecer_minutos_detenido(results, token, now_mx)
                except Exception:
                    logging.exception("No se pudo enriquecer detenciones")
            results = filtrar_contenido(results, reporte.get("contenido") or {})
            print(f"[{reporte['nombre']}] Recibidas={len(vehicles)} Incluidas={len(results)} Omitidas={len(excluidas)}")
            entregar_reporte(reporte, results, excluidas, now_mx, args.dry_run, args.canal)
        except Exception as error:
            logging.exception("Fallo el reporte %s", reporte["nombre"])
            errores.append(f"{reporte['nombre']}: {error}"); print(f"[ERROR] [{reporte['nombre']}] {error}")
    if errores:
        print("\nReportes con error:"); [print(f"- {error}") for error in errores]; return 1
    return 0


def crear_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--solo", action="append", help="Ejecuta solo este reporte; se puede repetir.")
    parser.add_argument(
        "--canal",
        choices=("google_chat", "correo"),
        help="Fuerza un canal configurado, aunque su entrega este inactiva en el JSON.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Previsualiza sin enviar.")
    parser.add_argument("--listar-etiquetas", action="store_true", help="Lista ID y nombre de etiquetas.")
    parser.add_argument("--sincronizar-etiquetas", action="store_true", help="Genera el catalogo local de padres e hijos.")
    parser.add_argument("--tag-page-size", type=int, default=10, help="Tamano de pagina para descargar tags (1-512).")
    return parser


def main() -> None:
    try:
        sys.exit(ejecutar(crear_parser().parse_args()))
    except Exception as error:
        logging.exception("Error fatal"); print(f"[ERROR] {error}"); sys.exit(1)


if __name__ == "__main__":
    main()
