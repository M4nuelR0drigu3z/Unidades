from datetime import datetime
from io import BytesIO
from pathlib import Path
import unittest

from openpyxl import load_workbook
import pytz

import EnvioMain as envio


class EnvioMainTests(unittest.TestCase):
    def test_reporte_ec02_usa_solo_las_seis_etiquetas_sayer_directas(self):
        config = envio.cargar_configuracion(Path("config/reportes.json"))
        reporte = next(x for x in config["reportes"] if x["nombre"] == "Reporte EC-02")
        self.assertEqual(reporte["tipo_filtro_etiqueta"], "tagIds")
        self.assertEqual(
            reporte["etiquetas"],
            [
                "Sayer Full",
                "Sayer Patios y T.",
                "Sayer Pipas",
                "Sayer Sencillo",
                "Sayer Thorton",
                "Sayer Vuelteros",
            ],
        )

    def test_construye_jerarquia_de_padres_e_hijos(self):
        tags = [
            {"id": "1", "nombre": "Padre", "parentTagId": ""},
            {"id": "2", "nombre": "Hijo", "parentTagId": "1"},
            {"id": "3", "nombre": "Nieto", "parentTagId": "2"},
        ]
        catalogo = envio.construir_jerarquia_etiquetas(tags)
        self.assertEqual(catalogo["totalEtiquetas"], 3)
        self.assertEqual(catalogo["totalEtiquetasPadre"], 2)
        self.assertEqual(catalogo["jerarquia"][0]["hijos"][0]["nombre"], "Hijo")
        self.assertEqual(
            catalogo["jerarquia"][0]["hijos"][0]["hijos"][0]["nombre"], "Nieto"
        )

    def test_perfil_ec02_sobrescribe_exclusion_global(self):
        base = {"excluir_unidades_en_geocerca": True, "geocercas_especiales_ids": ["1"]}
        perfiles = {
            "EC-02": {
                "excluir_unidades_en_geocerca": False,
                "geocercas_especiales_ids": [],
            }
        }
        filtros = envio.combinar_filtros(base, {"perfil_filtros": "EC-02"}, perfiles)
        self.assertFalse(filtros["excluir_unidades_en_geocerca"])
        self.assertEqual(filtros["geocercas_especiales_ids"], [])

    def test_busqueda_remota_por_nombre_usa_external_id_y_cache(self):
        class Response:
            status_code = 200

            def raise_for_status(self):
                return None

            def json(self):
                return {"data": {"id": "55", "name": "Sayer Full"}}

        class Session:
            def __init__(self):
                self.urls = []

            def get(self, url, timeout):
                self.urls.append(url)
                return Response()

        session, cache = Session(), {}
        primero = envio.resolver_etiquetas_samsara(session, ["Sayer Full"], cache)
        segundo = envio.resolver_etiquetas_samsara(session, ["sayer full"], cache)
        self.assertEqual(primero, {"Sayer Full": "55"})
        self.assertEqual(segundo, {"sayer full": "55"})
        self.assertEqual(len(session.urls), 1)
        self.assertIn("samsara.name:Sayer%20Full", session.urls[0])

    def test_paginacion_conserva_parametros_y_cursor(self):
        class Response:
            status_code = 200
            url = "https://api.samsara.com/tags"

            def __init__(self, payload):
                self.payload = payload

            def raise_for_status(self):
                return None

            def json(self):
                return self.payload

        class Session:
            def __init__(self):
                self.calls = []

            def get(self, url, params, timeout):
                self.calls.append((url, params, timeout))
                if len(self.calls) == 1:
                    return Response({"data": [{"id": "1"}], "pagination": {"hasNextPage": True, "endCursor": "abc"}})
                return Response({"data": [{"id": "2"}], "pagination": {"hasNextPage": False}})

        session = Session()
        data = envio.obtener_paginas(session, "/tags", {"limit": 10})
        self.assertEqual([x["id"] for x in data], ["1", "2"])
        self.assertEqual(session.calls[0][1], {"limit": 10})
        self.assertEqual(session.calls[1][1], {"limit": 10, "after": "abc"})

    def test_resuelve_etiquetas_sin_importar_mayusculas_o_acentos(self):
        tags = [{"id": "10", "name": "Sayer Full"}, {"id": "20", "name": "Sáyer Pipas"}]
        resultado = envio.resolver_etiquetas(["sayer full", "SAYER PIPAS"], tags)
        self.assertEqual(resultado, {"sayer full": "10", "SAYER PIPAS": "20"})

    def test_etiqueta_faltante_falla_claramente(self):
        with self.assertRaisesRegex(envio.ConfiguracionError, "no encontradas"):
            envio.resolver_etiquetas(["No existe"], [])

    def test_geocerca_es_configurable_por_reporte(self):
        tz = pytz.timezone("America/Mexico_City")
        now = tz.localize(datetime(2026, 8, 19, 12, 0))
        vehicles = [{"id": "1", "name": "2239", "gps": {
            "time": now.isoformat(), "speedMilesPerHour": 0, "isEcuSpeed": True,
            "address": {"id": "99", "name": "Patio"},
        }}]
        incluidos, omitidos = envio.procesar_vehiculos(vehicles, now, {"gps_max_minutos": 60}, {"99": "Patio"})
        self.assertEqual(incluidos, [])
        self.assertEqual(omitidos[0]["Motivo"], "PATIO/GEOCERCA EXCLUIDA")
        self.assertIn("Coordenadas", omitidos[0])
        incluidos, omitidos = envio.procesar_vehiculos(vehicles, now, {"gps_max_minutos": 60}, {})
        self.assertEqual(omitidos, [])
        self.assertEqual(incluidos[0]["Estatus"], "DETENIDO")

    def test_incluir_todas_conserva_gps_viejo_y_speed_cero_sin_ecu(self):
        tz = pytz.timezone("America/Mexico_City")
        now = tz.localize(datetime(2026, 8, 19, 12, 0))
        vehicles = [{"id": "1", "name": "100", "gps": {
            "time": "2026-08-18T10:00:00-06:00", "speedMilesPerHour": 0,
            "isEcuSpeed": False, "latitude": 19.1, "longitude": -99.2,
            "reverseGeo": {"formattedLocation": "Ubicacion anterior"},
        }}]
        filtros = {
            "gps_max_minutos": 60,
            "excluir_speed_cero_sin_ecu": True,
            "incluir_todas_las_unidades": True,
        }
        incluidos, omitidos = envio.procesar_vehiculos(vehicles, now, filtros, {})
        self.assertEqual(len(incluidos), 1)
        self.assertEqual(omitidos, [])
        self.assertEqual(incluidos[0]["Coordenadas"], "19.1,-99.2")

    def test_operador_samsara_usa_asignacion_mas_reciente(self):
        class Session:
            def get(self, url, params, timeout):
                class Response:
                    status_code = 200
                    url = "https://api.samsara.com/fleet/driver-vehicle-assignments"

                    def raise_for_status(self):
                        return None

                    def json(self):
                        return {"data": [
                            {"startTime": "2026-08-20T10:00:00Z", "isPassenger": False,
                             "vehicle": {"id": "1"}, "driver": {"id": "10", "name": "Anterior"}},
                            {"startTime": "2026-08-20T11:00:00Z", "isPassenger": False,
                             "vehicle": {"id": "1"}, "driver": {"id": "20", "name": "Actual"}},
                        ], "pagination": {"hasNextPage": False}}
                return Response()

        now = datetime(2026, 8, 20, 12, 0, tzinfo=pytz.UTC)
        datos = [{"Unidad": "100", "SamsaraVehicleId": "1"}]
        resultado = envio.enriquecer_operadores_samsara(Session(), datos, now)
        self.assertEqual(resultado[0]["Operador"], "Actual")
        self.assertEqual(resultado[0]["ID Operador"], "20")

    def test_excel_contiene_tres_hojas(self):
        now = datetime(2026, 8, 19, 12, 0, tzinfo=pytz.UTC)
        datos = [{"Unidad": "100", "Estatus": "RUTA", "Fecha GPS": now}]
        omitidas = [{"Unidad": "200", "Motivo": "GPS VIEJO"}]
        contenido = {"columnas": ["Unidad", "Estatus", "Fecha GPS"], "incluir_omitidas_excel": True}
        archivo = envio.crear_excel_reporte("Prueba", datos, omitidas, now, contenido)
        wb = load_workbook(BytesIO(archivo), read_only=True, data_only=True)
        self.assertEqual(wb.sheetnames, ["Resumen", "Unidades", "Omitidas"])
        self.assertEqual(wb["Unidades"]["A6"].value, "100")
        self.assertEqual(wb["Omitidas"]["A2"].value, "200")
        wb.close()

    def test_excel_puede_concentrarse_en_una_sola_hoja(self):
        now = datetime(2026, 8, 20, 12, 0, tzinfo=pytz.UTC)
        datos = [
            {"Unidad": "100", "Estatus": "RUTA"},
            {"Unidad": "200", "Estatus": "DETENIDO"},
        ]
        contenido = {
            "columnas": ["Unidad", "Estatus"],
            "resumen_estados_excel": True,
            "incluir_resumen_excel": False,
            "incluir_omitidas_excel": False,
        }
        archivo = envio.crear_excel_reporte("EC-02", datos, [], now, contenido)
        wb = load_workbook(BytesIO(archivo), read_only=True, data_only=True)
        self.assertEqual(wb.sheetnames, ["Unidades"])
        self.assertEqual(wb["Unidades"]["A6"].value, "100")
        self.assertEqual(wb["Unidades"]["B3"].value, 2)
        self.assertEqual(wb["Unidades"]["D3"].value, 1)
        self.assertEqual(wb["Unidades"]["F3"].value, 1)
        wb.close()

    def test_contenido_ordena_detenidos_antes_de_ruta(self):
        datos = [
            {"Unidad": "2", "Estatus": "RUTA"},
            {"Unidad": "10", "Estatus": "DETENIDO"},
            {"Unidad": "1", "Estatus": "DETENIDO"},
        ]
        resultado = envio.filtrar_contenido(
            datos, {"orden_estados": ["DETENIDO", "RUTA"]}
        )
        self.assertEqual([x["Unidad"] for x in resultado], ["1", "10", "2"])

    def test_divide_chat_sin_perder_lineas(self):
        self.assertEqual(envio.dividir_mensaje("uno\ndos\ntres", 8), ["uno\ndos", "tres"])

    def test_google_chat_incluye_coordenadas(self):
        tz = pytz.timezone("America/Mexico_City")
        now = tz.localize(datetime(2026, 8, 19, 12, 0))
        texto = envio.construir_reporte_google(
            [{"Unidad": "100", "Estatus": "RUTA", "Coordenadas": "19.1,-99.2"}],
            now,
            "Prueba",
        )
        self.assertIn("COORDENADAS", texto)
        self.assertIn("19.1,-99.2", texto)

    def test_entregas_usan_activas_si_no_se_fuerza_canal(self):
        reporte = {
            "nombre": "EC-05",
            "entregas": [
                {"canal": "google_chat", "activo": True},
                {"canal": "correo", "activo": False},
            ],
        }
        entregas = envio.seleccionar_entregas(reporte)
        self.assertEqual([x["canal"] for x in entregas], ["google_chat"])

    def test_canal_forzado_selecciona_correo_aunque_este_inactivo(self):
        reporte = {
            "nombre": "EC-05",
            "entregas": [
                {"canal": "google_chat", "activo": True},
                {"canal": "correo", "activo": False},
            ],
        }
        entregas = envio.seleccionar_entregas(reporte, "correo")
        self.assertEqual([x["canal"] for x in entregas], ["correo"])

    def test_solo_puede_seleccionar_reporte_inactivo(self):
        config = {
            "reportes": [
                {"nombre": "Reporte EC-05", "activo": True},
                {"nombre": "Reporte EC-02", "activo": False},
            ]
        }
        seleccionados = envio.seleccionar_reportes(config, ["Reporte EC-02"])
        self.assertEqual([x["nombre"] for x in seleccionados], ["Reporte EC-02"])

    def test_sin_solo_conserva_unicamente_reportes_activos(self):
        config = {
            "reportes": [
                {"nombre": "Reporte EC-05", "activo": True},
                {"nombre": "Reporte EC-02", "activo": False},
            ]
        }
        seleccionados = envio.seleccionar_reportes(config, None)
        self.assertEqual([x["nombre"] for x in seleccionados], ["Reporte EC-05"])


if __name__ == "__main__":
    unittest.main()
